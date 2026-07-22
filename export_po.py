#!/usr/bin/env python3
"""
export_po.py — Export PO restock data from KiotViet to clean Excel (.xlsx) files,
grouped by NPP (Nhà Phân Phối) according to brand mapping.
"""

import os
import re
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palette
COLOR_HEADER_BG = "1E293B"      # Dark slate header
COLOR_TABLE_HEAD = "2563EB"     # Royal Blue table header
COLOR_OUT_BG     = "FEE2E2"     # Light Red for 🔴 HẾT HÀNG
COLOR_OUT_TXT    = "991B1B"
COLOR_SOON_BG    = "FEF3C7"     # Light Yellow for 🟡 SẮP HẾT
COLOR_SOON_TXT   = "92400E"
COLOR_TOTAL_BG   = "F1F5F9"

font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
font_sub = Font(name="Segoe UI", size=10, italic=True, color="64748B")
font_head = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=10, bold=True)
font_regular = Font(name="Segoe UI", size=10)

font_out = Font(name="Segoe UI", size=10, bold=True, color=COLOR_OUT_TXT)
font_soon = Font(name="Segoe UI", size=10, bold=True, color=COLOR_SOON_TXT)

fill_title = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_head = PatternFill(start_color=COLOR_TABLE_HEAD, end_color=COLOR_TABLE_HEAD, fill_type="solid")
fill_out = PatternFill(start_color=COLOR_OUT_BG, end_color=COLOR_OUT_BG, fill_type="solid")
fill_soon = PatternFill(start_color=COLOR_SOON_BG, end_color=COLOR_SOON_BG, fill_type="solid")
fill_total = PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
top_thick_bottom_double = Border(
    top=Side(style='thin', color='475569'),
    bottom=Side(style='double', color='475569')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

# Brand List
BRANDS_LIST = ["AhaStyle", "Anank", "Anker", "Aulumu", "Baseus", "BMX", "Coteetci", "Divoom",
               "Elago", "HODA", "HyperWork", "Hyper", "IDMIX", "inateck", "Innostyle", "JCPAL",
               "Jinya", "JOWAY", "JRC", "LISEN", "Lofree", "Lucas", "Maxco", "Mipow", "Mocoll",
               "Nillkin", "NJOYNY", "Pitaka", "Satechi", "Sharge", "SKINARMA", "Thule", "Tomtoc",
               "Torras", "Ulanzi", "Urr", "WIWU", "Zagg", "UNIQ"]
_BRANDS_SORTED = sorted(BRANDS_LIST, key=len, reverse=True)

# NPP Mapping
NPP_MAP = {
    # HD: Thule, Innostyle, Tomtoc, BMX, Hyper, Mipow
    "thule": "HD",
    "innostyle": "HD",
    "tomtoc": "HD",
    "bmx": "HD",
    "hyper": "HD",
    "hyperwork": "HD",
    "mipow": "HD",

    # Viettel Distribution: Anker
    "anker": "Viettel Distribution",

    # iFuture: Aulumu, Lisen
    "aulumu": "iFuture",
    "lisen": "iFuture",

    # ACE: JRC, Inateck
    "jrc": "ACE",
    "inateck": "ACE",

    # Senco: JCPAL, IDMIX, Mocoll
    "jcpal": "Senco",
    "idmix": "Senco",
    "mocoll": "Senco",

    # Tuấn Anh: Hoda, Joway, Maxco, Nillkin, WIWU
    "hoda": "Tuấn Anh",
    "joway": "Tuấn Anh",
    "maxco": "Tuấn Anh",
    "nillkin": "Tuấn Anh",
    "wiwu": "Tuấn Anh",

    # DTR: Pitaka, Zagg
    "pitaka": "DTR",
    "zagg": "DTR",

    # TLC: Uniq, Skinarma
    "uniq": "TLC",
    "skinarma": "TLC",

    # Huy Linh: Ulanzi
    "ulanzi": "Huy Linh",

    # Lucas: Lucas
    "lucas": "Lucas",

    # StreamCast: Satechi, Sharge
    "satechi": "StreamCast",
    "sharge": "StreamCast",
}

def get_brand(name):
    low = (name or "").lower()
    for b in _BRANDS_SORTED:
        if b.lower() in low:
            return b
    return "Khác"

def get_npp(brand):
    b_low = (brand or "").lower()
    return NPP_MAP.get(b_low, "Khác")

def create_po_sheet(ws, sheet_title, items, cover_days=30, is_master=False):
    ws.views.sheetView[0].showGridLines = True

    # Column count
    col_count = 9 if is_master else 8
    last_col_letter = get_column_letter(col_count)

    # 1. Title Block
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell_title = ws["A1"]
    cell_title.value = f"ĐƠN ĐẶT HÀNG NHẬP KHO (PURCHASE ORDER) — {sheet_title.upper()}"
    cell_title.font = font_title
    cell_title.fill = fill_title
    cell_title.alignment = align_center
    ws.row_dimensions[1].height = 36

    # Sub info
    today_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells(f"A2:{last_col_letter}2")
    cell_sub = ws["A2"]
    cell_sub.value = f"Ngày khởi tạo: {today_str} | Target dự trữ: {cover_days} ngày | Đơn vị: Combomacbook / Lucas.vn"
    cell_sub.font = font_sub
    cell_sub.alignment = align_center
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10

    # Table Headers
    if is_master:
        headers = [
            ("STT", 6, align_center),
            ("NPP (Nhà Phân Phối)", 20, align_left),
            ("Thương Hiệu", 15, align_left),
            ("Mã SP (SKU)", 16, align_left),
            ("Tên Sản Phẩm", 42, align_left),
            ("Tồn Kho", 10, align_right),
            ("Bán/Ngày", 10, align_right),
            ("SL Nhập Gợi Ý", 14, align_right),
            ("Trạng Thái", 14, align_center)
        ]
    else:
        headers = [
            ("STT", 6, align_center),
            ("Thương Hiệu", 15, align_left),
            ("Mã SP (SKU)", 16, align_left),
            ("Tên Sản Phẩm", 45, align_left),
            ("Tồn Kho", 10, align_right),
            ("Bán/Ngày", 10, align_right),
            ("SL Nhập Gợi Ý", 14, align_right),
            ("Trạng Thái", 14, align_center)
        ]

    ws.row_dimensions[4].height = 26
    for col_idx, (h_name, _, h_align) in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h_name)
        c.font = font_head
        c.fill = fill_head
        c.alignment = h_align
        c.border = thin_border

    # Data Rows
    start_row = 5
    for idx, item in enumerate(items, start=1):
        row_idx = start_row + idx - 1
        ws.row_dimensions[row_idx].height = 22

        is_out = item.get("oh", 0) <= 0
        st_text = "🔴 HẾT HÀNG" if is_out else f"🟡 Còn {item.get('dleft', 0):.0f} ngày"
        row_font = font_out if is_out else font_soon
        row_fill = fill_out if is_out else fill_soon

        if is_master:
            values = [
                (idx, align_center, None),
                (item.get("npp", "Khác"), align_left, None),
                (item.get("brand", "Khác"), align_left, None),
                (item.get("code", ""), align_left, None),
                (item.get("name", ""), align_left, None),
                (item.get("oh", 0), align_right, "#,##0"),
                (item.get("vel", 0.0), align_right, "0.0"),
                (item.get("need", 0), align_right, "#,##0"),
                (st_text, align_center, None)
            ]
        else:
            values = [
                (idx, align_center, None),
                (item.get("brand", "Khác"), align_left, None),
                (item.get("code", ""), align_left, None),
                (item.get("name", ""), align_left, None),
                (item.get("oh", 0), align_right, "#,##0"),
                (item.get("vel", 0.0), align_right, "0.0"),
                (item.get("need", 0), align_right, "#,##0"),
                (st_text, align_center, None)
            ]

        status_col = 9 if is_master else 8
        need_col = 8 if is_master else 7

        for col_idx, (val, alignment, num_fmt) in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = font_bold if col_idx in (1, need_col) else font_regular
            c.alignment = alignment
            c.border = thin_border
            if num_fmt:
                c.number_format = num_fmt
            if col_idx == status_col: # Highlight status cell
                c.font = row_font
                c.fill = row_fill

    # Total Row
    tot_row = start_row + len(items)
    ws.row_dimensions[tot_row].height = 25

    label_col_span = 5 if is_master else 4
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=label_col_span)
    cell_tot_label = ws.cell(row=tot_row, column=1, value=f"TỔNG CỘNG ({len(items)} SẢN PHẨM)")
    cell_tot_label.font = font_bold
    cell_tot_label.alignment = align_right

    for c_idx in range(1, col_count + 1):
        c = ws.cell(row=tot_row, column=c_idx)
        c.fill = fill_total
        c.border = top_thick_bottom_double

    # Sum formulas for Stock & Needed PO Qty
    stock_col_letter = "F" if is_master else "E"
    need_col_letter = "H" if is_master else "G"
    stock_col_idx = 6 if is_master else 5
    need_col_idx = 8 if is_master else 7

    c_oh = ws.cell(row=tot_row, column=stock_col_idx, value=f"=SUM({stock_col_letter}{start_row}:{stock_col_letter}{tot_row-1})")
    c_oh.font = font_bold
    c_oh.alignment = align_right
    c_oh.number_format = "#,##0"

    c_need = ws.cell(row=tot_row, column=need_col_idx, value=f"=SUM({need_col_letter}{start_row}:{need_col_letter}{tot_row-1})")
    c_need.font = font_bold
    c_need.alignment = align_right
    c_need.number_format = "#,##0"

    # Set Column Widths
    for col_idx, (_, min_w, _) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min_w


def export_po_excel(out_items, soon_items, output_dir="po_exports", cover_days=30):
    """
    Xuất danh sách sản phẩm ra file Excel (.xlsx) chuẩn PO, nhóm theo NPP (Nhà Phân Phối).
    Trả về dict: {"master_file": path, "brand_files": {npp: path}}
    """
    os.makedirs(output_dir, exist_ok=True)
    today_stamp = datetime.now().strftime("%Y%m%d_%H%M")

    # Combined list with Brand and NPP
    all_items = []
    for r in out_items:
        b = r.get("brand") or get_brand(r.get("name"))
        n = get_npp(b)
        all_items.append(dict(r, crit=0, st="HẾT", brand=b, npp=n))
    for r in soon_items:
        b = r.get("brand") or get_brand(r.get("name"))
        n = get_npp(b)
        all_items.append(dict(r, crit=1, st=f"{r['dleft']:.0f}d", brand=b, npp=n))

    if not all_items:
        print("[!] Không có sản phẩm cần nhập -> Không tạo file Excel.")
        return None

    # Group by NPP
    npp_groups = defaultdict(list)
    for it in all_items:
        npp_groups[it["npp"]].append(it)

    # Sort items within NPP groups (Hết hàng trước, rồi theo vận tốc bán)
    for n in npp_groups:
        npp_groups[n].sort(key=lambda i: (i["crit"], -i["vel"]))

    # Sort all items for Master sheet (by NPP, then Hết hàng, then vel)
    all_items.sort(key=lambda i: (i["npp"], i["crit"], -i["vel"]))

    # 1. Create Master Workbook with multiple sheets
    wb_master = openpyxl.Workbook()
    ws_all = wb_master.active
    ws_all.title = "TẤT CẢ NPP"
    create_po_sheet(ws_all, "TỔNG HỢP CẢNH BÁO KHO THEO NPP", all_items, cover_days=cover_days, is_master=True)

    # Sort NPPs by number of urgent items
    sorted_npps = sorted(npp_groups.keys(), key=lambda n: (-sum(1 for i in npp_groups[n] if i["crit"] == 0), -len(npp_groups[n])))

    npp_files = {}

    for n in sorted_npps:
        items = npp_groups[n]
        clean_title = re.sub(r'[\\/*?:\[\]]', '', str(n))[:30] or "Khac"

        # Add sheet to master workbook
        ws_n = wb_master.create_sheet(title=clean_title)
        create_po_sheet(ws_n, f"NPP {n}", items, cover_days=cover_days, is_master=False)

        # 2. Create individual Workbook per NPP for direct sending
        wb_n = openpyxl.Workbook()
        ws_single = wb_n.active
        ws_single.title = clean_title
        create_po_sheet(ws_single, f"NPP {n}", items, cover_days=cover_days, is_master=False)

        file_n_path = os.path.join(output_dir, f"PO_NPP_{clean_title.replace(' ', '_')}_{today_stamp}.xlsx")
        wb_n.save(file_n_path)
        npp_files[n] = file_n_path

    master_path = os.path.join(output_dir, f"PO_TONG_HOP_{today_stamp}.xlsx")
    wb_master.save(master_path)

    print(f"[+] Đã xuất Master PO: {master_path}")
    print(f"[+] Đã xuất {len(npp_files)} file PO riêng theo Nhà Phân Phối (NPP) trong folder '{output_dir}/'")

    return {
        "master_file": master_path,
        "brand_files": npp_files
    }

if __name__ == "__main__":
    # Quick test mock data
    sample_out = [
        {"code": "TOMTOC01", "name": "Balo Tomtoc A42 MacBook 16 inch", "oh": 0, "vel": 1.5, "dleft": 0, "need": 45},
        {"code": "THULE02", "name": "Balo Thule Subterra 26L Black", "oh": 0, "vel": 0.8, "dleft": 0, "need": 24},
        {"code": "ANKER01", "name": "Sạc Dự Phòng Anker 622 MagGo", "oh": 0, "vel": 2.0, "dleft": 0, "need": 60},
        {"code": "ULANZI01", "name": "Giá Đỡ Điện Thoại Ulanzi MA02", "oh": 0, "vel": 1.2, "dleft": 0, "need": 36}
    ]
    sample_soon = [
        {"code": "WIWU03", "name": "Túi Chống Sốc WiWU Pilot Sleeve 14 inch", "oh": 3, "vel": 0.5, "dleft": 6.0, "need": 12},
        {"code": "UNIQ01", "name": "Case Ốp MacBook UNIQ Claro", "oh": 2, "vel": 0.4, "dleft": 5.0, "need": 10}
    ]
    export_po_excel(sample_out, sample_soon)
